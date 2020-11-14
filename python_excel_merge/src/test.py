from openpyxl import load_workbook
import openpyxl
import glob

MERGE_XLSX="./excel/merge.xlsx"

files = glob.glob("./excel/*")
for file in files:
    print(file)


mergeWb = openpyxl.load_workbook(MERGE_XLSX)

wb = openpyxl.load_workbook('./excel/テーブル定義書（商品詳細マスタ）.xlsx')
ws = wb.worksheets[0]
mergeWb.copy_worksheet(ws)

mergeWb.save(MERGE_XLSX)

#対象のエクセルファイル指定
# xls_file='test.xls'

# #openpyxlでエクセルファイル（ブック）とアクティブシートを読込
# wb=load_workbook(xls_file)
# ws=wb.active

# #アクティブシートをコピー
# ws_copy=wb.copy_worksheet(ws)

# #コピーされたシートのタイトルを変更する
# ws_copy.title='新シート名'

# #リストで量産するなら
# list=['シート1','シート2','シート3','シート4']
# for sheet_name in list:
#     ws_copy = wb.copy_worksheet(ws)
#     ws_copy.title=sheet_name

# #保存して終わり
# wb.save(xls_file)